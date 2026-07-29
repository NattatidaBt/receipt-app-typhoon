import base64
import csv
import inspect
import json
import math
import re
from io import StringIO

import cv2
import streamlit as st
import streamlit.components.v1 as components

from llm_engine import call_typhoon_llm
from ocr_engine import (
    clean_ocr_text,
    deskew_image,
    format_ocr_text,
    load_image_or_pdf,
    process_clahe,
    run_typhoon_ocr,
)

st.set_page_config(
    page_title="RecAipt - Receipt Verification",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── ประเภทเอกสารที่รองรับ ───────────────────────────────────────────────────
DOC_TYPES = ["ใบเสร็จทั่วไป", "ใบกำกับภาษีอย่างย่อ", "ใบกำกับภาษีแบบเต็มรูป"]
DOC_TYPE_DEFAULT = DOC_TYPES[0]

# คำ/วลีที่ LLM อาจส่งกลับมา → map ไปยัง 3 ตัวเลือกมาตรฐาน
_DOC_TYPE_MAP = {
    # ใบเสร็จทั่วไป
    "ใบเสร็จรับเงิน": "ใบเสร็จทั่วไป",
    "ใบเสร็จ": "ใบเสร็จทั่วไป",
    "receipt": "ใบเสร็จทั่วไป",
    "ใบเสร็จทั่วไป": "ใบเสร็จทั่วไป",
    # ใบกำกับภาษีอย่างย่อ
    "ใบกำกับภาษีอย่างย่อ": "ใบกำกับภาษีอย่างย่อ",
    "ใบกำกับย่อ": "ใบกำกับภาษีอย่างย่อ",
    "tax invoice (abbreviated)": "ใบกำกับภาษีอย่างย่อ",
    "abbreviated tax invoice": "ใบกำกับภาษีอย่างย่อ",
    # ใบกำกับภาษีแบบเต็มรูป
    "ใบกำกับภาษี": "ใบกำกับภาษีแบบเต็มรูป",
    "ใบกำกับภาษีแบบเต็มรูป": "ใบกำกับภาษีแบบเต็มรูป",
    "ใบกำกับภาษีเต็มรูปแบบ": "ใบกำกับภาษีแบบเต็มรูป",
    "tax invoice": "ใบกำกับภาษีแบบเต็มรูป",
    "full tax invoice": "ใบกำกับภาษีแบบเต็มรูป",
}


def normalize_doc_type(value):
    """แปลงค่าจาก OCR/LLM ให้ตรงกับ DOC_TYPES หนึ่งใน 3 ตัว"""
    if not value:
        return DOC_TYPE_DEFAULT
    v = str(value).strip()
    # ตรงตัว
    if v in DOC_TYPES:
        return v
    # map แบบ case-insensitive
    mapped = _DOC_TYPE_MAP.get(v) or _DOC_TYPE_MAP.get(v.lower())
    if mapped:
        return mapped
    # เดาจาก keyword
    vl = v.lower()
    if "เต็ม" in vl or "full" in vl:
        return "ใบกำกับภาษีแบบเต็มรูป"
    if "ย่อ" in vl or "abbr" in vl:
        return "ใบกำกับภาษีอย่างย่อ"
    if "กำกับ" in vl or "invoice" in vl:
        return "ใบกำกับภาษีแบบเต็มรูป"
    return DOC_TYPE_DEFAULT


# สตริงรวมสไตล์ CSS ทั้งหมดของหน้าจอ
CSS_STYLES = """
<style>
:root {
    --bg: #F4F1EA;
    --panel: #FFFDF8;
    --panel-2: #EEF5F3;
    --ink: #1F2A2A;
    --muted: #66736F;
    --line: #D8D1C3;
    --accent: #2F6F73;
    --accent-2: #285E62;
    --ok-bg: #e8f2e7;
    --ok-line: #9abe98;
    --ok-ink: #2c6330;
    --warn-bg: #fff3d8;
    --warn-line: #dfb65b;
    --warn-ink: #75520f;
    --bad-bg: #fae4e0;
    --bad-line: #d9958d;
    --bad-ink: #91382d;
}

header, footer, #MainMenu,
[data-testid="stToolbar"],
[data-testid="stDecoration"] {
    display: none !important;
}

.stApp {
    background: var(--bg);
    color: var(--ink);
}

.block-container {
    max-width: 100% !important;
    padding: 0 22px 18px !important;
    margin-top: 0 !important;
    padding-top: 0 !important;
}

/* ลบ Gap และ Margin ของ Streamlit Container ทุกระดับชั้น */
[data-testid="stMainBlockContainer"] {
    padding-top: 0px !important;
    margin-top: 0px !important;
    padding-bottom: 24px !important;
}

div[data-testid="stVerticalBlockBorderWrapper"]:has(div[data-testid="stHorizontalBlock"]) {
    margin-top: 0px !important;
    padding-top: 0px !important;
}

div[data-testid="stVerticalBlock"]:has(> div [data-testid="element-container"] .app-topbar) + div {
    margin-top: 0px !important;
    padding-top: 0px !important;
}

[data-testid="stHorizontalBlock"] {
    margin-top: 0px !important; 
    padding-top: 0px !important;
    margin-bottom: 8px !important; 
    gap: 16px !important; 
}

[data-testid="stColumn"] [data-testid="stVerticalBlock"] {
    gap: 10px !important;
}

div[data-testid="stIFrame"] {
    margin-bottom: 0px !important;
}

[data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {
    padding-top: 0px !important;
    margin-top: 0px !important;
}

[data-testid="stHorizontalBlock"] > [data-testid="stColumn"] > div {
    padding-top: 0px !important;
    margin-top: 0px !important;
}

[data-testid="stVerticalBlock"] > [data-testid="element-container"] {
    margin-bottom: 0 !important;
    gap: 0 !important;
}

/* คงระยะห่างใต้ Topbar ไว้ ไม่ให้ถูกรีเซ็ตเป็น 0 */
[data-testid="stVerticalBlock"] > [data-testid="element-container"]:has(.app-topbar) {
    margin-bottom: 24px !important;
}

/* ปิดพื้นที่ว่างจาก iframe ของ Feedback Modal ที่ซ่อนไว้ท้ายหน้า */
div[data-testid="element-container"]:has(div[data-testid="stIFrame"]):not([data-testid="stColumn"] *) {
    display: none !important;
    height: 0 !important;
    min-height: 0 !important;
    margin: 0 !important;
    padding: 0 !important;
}

section.main, .main {
    padding-bottom: 0px !important;
}

h1, h2, h3, h4, p, label, span, div {
    letter-spacing: 0 !important;
}

.app-topbar {
    position: sticky;
    top: 0;
    z-index: 20;
    margin: 0 -22px 32px -22px;
    padding: 18px 28px 16px;
    background: rgba(234, 246, 251, 0.94);
    backdrop-filter: blur(6px);
    border-bottom: 1px solid var(--line);
}

.pane {
    background: var(--panel);
    border: none;
    border-radius: 10px;
    margin-top: 0px !important;
    box-shadow: 0 3px 14px rgba(15, 92, 112, 0.14);
}

.pane-head {
    padding: 18px 20px;
    background: var(--accent);
    border-radius: 10px;
}

.pane-title {
    color: #ffffff;
    font-size: 1.14rem;
    font-weight: 780;
    line-height: 1.35;
}

.pane-note {
    color: rgba(255, 255, 255, 0.86);
    font-size: 0.92rem;
    line-height: 1.5;
    margin-top: 10px;
}

/* รวบกล่องสไลเดอร์ซูมกับแถวหัวข้อให้เกาะสนิทกันเป็นแผ่นเดียว */
.zoom-control-row {
    background: var(--panel);
    border-left: 1px solid var(--line);
    border-right: 1px solid var(--line);
    padding: 4px 16px 0px 16px !important;
    margin-top: 25px !important;
}

.zoom-control-row [data-testid="stSlider"] label,
.zoom-control-row [data-testid="stSelectbox"] label {
    display: none !important;
}

.zoom-control-row [data-testid="element-container"] {
    margin-bottom: 0px !important;
    padding-top: 0px !important;
    margin-top: 0px !important;
}

/* ล็อกความสูงกล่องแสดงภาพฝั่งซ้ายให้เกาะยืดตัวยาวอิงตามคอลัมน์ฟอร์มขวาเพื่อปิดรอยแหว่งครีม */
.image-scroll {
    height: calc(100vh - 260px) !important;
    min-height: 520px !important;
    max-height: 780px !important;
    overflow: auto;
    padding: 14px 14px 28px 14px;
    background:
        linear-gradient(45deg, #E7E1D4 25%, transparent 25%),
        linear-gradient(-45deg, #E7E1D4 25%, transparent 25%),
        linear-gradient(45deg, transparent 75%, #E7E1D4 75%),
        linear-gradient(-45deg, transparent 75%, #E7E1D4 75%);
    background-size: 18px 18px;
    background-position: 0 0, 0 9px, 9px -9px, -9px 0;
    border: 1px solid var(--line);
    border-radius: 0 0 8px 8px;
    margin-top: 0px !important;
    margin-bottom: 16px !important;
}

.receipt-image {
    display: block;
    height: auto;
    margin: 0 auto;
    border: 1px solid var(--line);
    background: #fff;
    box-shadow: 0 10px 28px rgba(31, 41, 55, 0.12);
}

.topbar-grid {
    display: flex;
    align-items: center;
    justify-content: space-between;
    width: 100%;
    gap: 0;
}

.topbar-left {
    display: flex;
    align-items: center;
    gap: 12px;
    flex-wrap: wrap;
    flex: initial;
    min-width: 0;
}

.topbar-divider {
    width: 1px;
    height: 20px;
    background: var(--line);
    flex-shrink: 0;
    margin: 0 12px;
}

.brand {
    color: var(--ink);
    font-size: 1.1rem;
    font-weight: 720;
    line-height: 1;
    white-space: nowrap;
}

.status-row {
    display: flex;
    gap: 6px;
    flex-wrap: wrap;
    align-items: center;
}

.badge {
    display: inline-flex;
    align-items: center;
    min-height: 24px;
    padding: 3px 9px;
    border-radius: 999px;
    font-size: 0.8rem;
    font-weight: 600;
    white-space: nowrap;
}

.badge-ok {
    color: #1a4d1e;
    background: #d6efd8;
    border: 1px solid #8fc492;
}

.badge-warn {
    color: #5c3d08;
    background: #fdefc8;
    border: 1px solid #d9a93a;
}

.badge-bad {
    color: #6b1f1a;
    background: #fce4e0;
    border: 1px solid #d9958d;
}

.upload-wrap {
    max-width: 960px;
    margin: 0 auto;
    padding: 0 10px;
}

.upload-page-shell {
    width: min(960px, calc(100vw - 48px));
    margin: clamp(28px, 8vh, 76px) auto 0;
}

.upload-title {
    color: var(--ink);
    font-size: clamp(1.65rem, 2.2vw, 2.25rem);
    font-weight: 760;
    line-height: 1.25;
    text-align: center;
    margin: 0 0 8px;
}

.upload-subtitle {
    color: var(--muted);
    font-size: 1.08rem;
    line-height: 1.65;
    text-align: center;
    margin: 0 auto 22px;
    max-width: 760px;
}

.upload-tip {
    max-width: 760px;
    margin: 28px auto 0;
}

/* ─── FILE UPLOADER: จัดกึ่งกลาง X,Y และปุ่มสีขาว ─── */
[data-testid="stFileUploader"] {
    max-width: 820px;
    margin: 0 auto !important;
}

[data-testid="stFileUploaderDropzone"] {
    background: var(--panel) !important;
    border: 2px dashed #8fc7d9 !important;
    border-radius: 8px !important;
    min-height: 168px !important;
    padding: 26px 30px !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
}

[data-testid="stFileUploaderDropzone"] > div {
    display: flex !important;
    flex-direction: row !important;
    align-items: center !important;
    justify-content: center !important;
    width: 100% !important;
    gap: 24px !important;
}

[data-testid="stFileUploaderDropzone"]:hover {
    border-color: var(--accent) !important;
    background: #f5fcfe !important;
}

[data-testid="stFileUploaderDropzoneInstructions"] {
    display: flex !important;
    flex-direction: column !important;
    align-items: flex-start !important;
    justify-content: center !important;
}

[data-testid="stFileUploaderDropzoneInstructions"] span {
    color: var(--ink) !important;
    font-size: 1.05rem !important;
    line-height: 1.5 !important;
}

[data-testid="stFileUploaderDropzoneInstructions"] small {
    color: var(--muted) !important;
    font-size: 0.96rem !important;
}

/* Override ปุ่ม Browse files สีดำบน Streamlit Cloud */
[data-testid="stFileUploaderDropzone"] button,
[data-testid="stFileUploader"] button,
[data-testid="stFileUploader"] section button {
    background-color: var(--panel) !important;
    background: var(--panel) !important;
    color: var(--ink) !important;
    border: 1px solid var(--line) !important;
    border-radius: 8px !important;
    box-shadow: none !important;
    min-height: 40px !important;
    padding: 6px 18px !important;
    font-size: 0.98rem !important;
    font-weight: 600 !important;
    cursor: pointer !important;
    flex-shrink: 0 !important;
}

[data-testid="stFileUploaderDropzone"] button:hover,
[data-testid="stFileUploader"] button:hover,
[data-testid="stFileUploader"] section button:hover {
    background-color: var(--panel-2) !important;
    background: var(--panel-2) !important;
    border-color: var(--accent) !important;
    color: var(--accent-2) !important;
}

/* ชื่อไฟล์ที่เลือกแล้วบน Streamlit Cloud */
[data-testid="stFileUploaderFile"] *,
[data-testid="stFileUploaderFileName"],
[data-testid="stFileUploaderFileSize"] {
    color: var(--ink) !important;
    opacity: 1 !important;
}

[data-testid="stFileUploaderFile"] {
    background: transparent !important;
}

[data-testid="stFileUploaderDeleteBtn"] button,
[data-testid="stFileUploaderDeleteBtn"],
[data-testid="stFileUploaderFile"] button {
    background: var(--panel) !important;
    color: var(--ink) !important;
    border: 1px solid var(--line) !important;
    border-radius: 8px !important;
}
/* ─────────────────────────────────────────────────── */

.soft-callout {
    padding: 12px 14px;
    border-radius: 8px;
    background: var(--panel-2);
    border: 1px solid var(--line);
    color: var(--accent-2);
    font-size: 0.97rem;
    line-height: 1.55;
    margin: 10px 0 14px;
}

.section-band {
    margin-top: 12px;
    margin-bottom: 10px;
    padding: 9px 12px;
    border-left: 5px solid var(--accent);
    background: #cdeaf5;
    color: var(--accent-2);
    font-size: 1.02rem;
    font-weight: 780;
}

.metric-grid {
    display: grid;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    gap: 10px;
    margin: 8px 0 12px;
}

.metric {
    padding: 12px;
    border: 1px solid var(--line);
    border-radius: 8px;
    background: var(--panel);
}

.metric-label {
    color: var(--muted);
    font-size: 0.9rem;
    line-height: 1.4;
}

.metric-value {
    color: var(--ink);
    font-size: 1.18rem;
    font-weight: 760;
    line-height: 1.3;
    margin-top: 4px;
    overflow-wrap: anywhere;
}

.review-list {
    margin: 8px 0 12px;
    padding: 0;
    list-style: none;
}

.review-list li {
    margin: 6px 0;
    padding: 9px 11px;
    border-radius: 8px;
    font-size: 0.96rem;
    line-height: 1.5;
}

.needs-check {
    background: var(--warn-bg);
    border: 1px solid var(--warn-line);
    color: var(--warn-ink);
}

.looks-ok {
    background: var(--ok-bg);
    border: 1px solid var(--ok-line);
    color: var(--ok-ink);
}

.stTextInput input,
.stTextArea textarea,
.stNumberInput input,
.stDateInput input,
.stSelectbox div[data-baseweb="select"] > div {
    min-height: 44px !important;
    font-size: 1.02rem !important;
    line-height: 1.55 !important;
    color: var(--ink) !important;
    background: var(--panel) !important;
    border-color: var(--line) !important;
    border-radius: 8px !important;
}

.stTextArea textarea {
    min-height: 92px !important;
}

.stTextInput label,
.stTextArea label,
.stNumberInput label,
.stSelectbox label,
.stDataFrame label {
    color: var(--ink) !important;
    font-size: 0.98rem !important;
    font-weight: 680 !important;
    line-height: 1.45 !important;
}

.stTextInput input:focus,
.stTextArea textarea:focus,
.stNumberInput input:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 3px rgba(47, 111, 115, 0.18) !important;
}

div.stButton > button,
div.stDownloadButton > button,
div.stFormSubmitButton > button {
    min-height: 44px !important;
    border-radius: 8px !important;
    font-size: 1rem !important;
    font-weight: 720 !important;
    border: 1px solid var(--line) !important;
    background: var(--panel) !important;
    color: var(--ink) !important;
}

div.stButton > button:hover,
div.stDownloadButton > button:hover,
div.stFormSubmitButton > button:hover {
    border-color: var(--accent) !important;
    color: var(--accent-2) !important;
    background: var(--panel-2) !important;
}

div.stButton > button[kind="primary"],
div.stFormSubmitButton > button[kind="primary"] {
    background: var(--accent) !important;
    color: #ffffff !important;
    border-color: var(--accent) !important;
}

div.stButton > button[kind="primary"]:hover,
div.stFormSubmitButton > button[kind="primary"]:hover {
    background: var(--accent-2) !important;
    color: #ffffff !important;
}

div.stPageLink > a,
div[data-testid="stPageLink"] > a {
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    width: 100% !important;
    min-height: 44px !important;
    border-radius: 8px !important;
    font-size: 1rem !important;
    font-weight: 720 !important;
    border: 1px solid var(--line) !important;
    background: var(--panel) !important;
    color: var(--ink) !important;
    text-decoration: none !important;
    padding: 0 16px !important;
    box-sizing: border-box !important;
    gap: 6px !important;
    transition: background 0.15s, border-color 0.15s, color 0.15s !important;
}

div.stPageLink > a:hover,
div[data-testid="stPageLink"] > a:hover {
    border-color: var(--accent) !important;
    color: var(--accent-2) !important;
    background: var(--panel-2) !important;
    text-decoration: none !important;
}

.stNumberInput button,
.stNumberInput [data-testid="stNumberInputStepDown"],
.stNumberInput [data-testid="stNumberInputStepUp"] {
    background: var(--panel) !important;
    color: var(--ink) !important;
    border: 1px solid var(--line) !important;
}

.stNumberInput button:hover,
.stNumberInput [data-testid="stNumberInputStepDown"]:hover,
.stNumberInput [data-testid="stNumberInputStepUp"]:hover {
    background: var(--panel-2) !important;
    color: var(--accent-2) !important;
    border-color: var(--accent) !important;
}

.stDateInput input,
.stDateInput input::placeholder,
.stTextInput input::placeholder {
    color: var(--ink) !important;
    opacity: 1 !important;
}

[data-testid="stDataFrame"] {
    border: 1px solid var(--line);
    border-radius: 8px;
    overflow: hidden;
    background: var(--panel) !important;
    color: var(--ink) !important;
}

[data-testid="stDataFrame"] * {
    color: var(--ink) !important;
}

[data-testid="stDataFrame"] div,
[data-testid="stDataFrame"] span {
    background-color: transparent;
}

[data-testid="stDataFrame"] button {
    background: var(--panel) !important;
    color: var(--ink) !important;
    border-color: var(--line) !important;
}

[data-testid="stAlert"] *,
[data-testid="stException"] * {
    color: var(--ink) !important;
}

.action-pad {
    padding: 12px;
    margin-top: 12px;
    background: rgba(255, 255, 255, 0.96);
    border: 1px solid var(--line);
    border-radius: 8px;
    box-shadow: 0 -8px 24px rgba(31, 41, 55, 0.08);
}

.kbd {
    display: inline-block;
    padding: 1px 7px;
    border: 1px solid var(--line);
    border-radius: 5px;
    background: var(--panel);
    color: var(--ink);
    font-family: ui-monospace, SFMono-Regular, Consolas, monospace;
    font-size: 0.88rem;
}

.tiny-note {
    color: var(--muted);
    font-size: 0.92rem;
    line-height: 1.55;
}

.fb-overlay {
    position: fixed;
    inset: 0;
    z-index: 9999;
    background: rgba(15, 40, 48, 0.48);
    display: flex;
    align-items: center;
    justify-content: center;
}

.fb-modal {
    background: var(--panel);
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 28px 32px 24px;
    width: min(520px, 92vw);
    box-shadow: 0 24px 64px rgba(15, 40, 48, 0.22);
}

.fb-title {
    color: var(--ink);
    font-size: 1.18rem;
    font-weight: 760;
    line-height: 1.35;
    margin-bottom: 6px;
}

.fb-sub {
    color: var(--muted);
    font-size: 0.96rem;
    line-height: 1.55;
    margin-bottom: 18px;
}

.fb-chip-row {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-bottom: 16px;
}

.fb-chip {
    padding: 6px 14px;
    border-radius: 999px;
    border: 1px solid var(--line);
    background: var(--panel);
    color: var(--ink);
    font-size: 0.94rem;
    cursor: pointer;
    user-select: none;
    transition: background 0.15s, border-color 0.15s;
}

.fb-chip:hover {
    border-color: var(--accent);
    background: var(--panel-2);
    color: var(--accent-2);
}

.fb-chip.selected {
    background: var(--accent);
    border-color: var(--accent);
    color: #fff;
}

.fb-textarea {
    width: 100%;
    min-height: 84px;
    padding: 10px 12px;
    border: 1px solid var(--line);
    border-radius: 8px;
    background: var(--panel);
    color: var(--ink);
    font-size: 1rem;
    line-height: 1.55;
    resize: vertical;
    box-sizing: border-box;
    font-family: inherit;
    margin-bottom: 16px;
}

.fb-textarea:focus {
    outline: none;
    border-color: var(--accent);
    box-shadow: 0 0 0 3px rgba(47, 111, 115, 0.18) !important;
}

.fb-btn-row {
    display: flex;
    gap: 10px;
    justify-content: flex-end;
}

.fb-btn {
    min-height: 40px;
    padding: 6px 20px;
    border-radius: 8px;
    border: 1px solid var(--line);
    background: var(--panel);
    color: var(--ink);
    font-size: 0.98rem;
    font-weight: 720;
    cursor: pointer;
    transition: background 0.15s, border-color 0.15s;
}

.fb-btn:hover {
    border-color: var(--accent);
    background: var(--panel-2);
}

.fb-btn.primary {
    background: var(--accent);
    border-color: var(--accent);
    color: #fff;
}

.fb-btn.primary:hover {
    background: var(--accent-2);
}

.fb-star-row {
    display: flex;
    gap: 6px;
    margin-bottom: 16px;
}

.fb-star {
    font-size: 1.6rem;
    cursor: pointer;
    color: #bcdbe4;
    transition: color 0.12s;
    user-select: none;
}

.fb-star.lit {
    color: #e8a820;
}

.ocr-expander-spacer {
    height: 14px;
}

@media (max-width: 980px) {
    .topbar-grid,
    .split-shell,
    .metric-grid {
        grid-template-columns: 1fr;
    }

    .status-row {
        justify-content: flex-start;
    }

    .image-scroll {
        height: 58vh;
        min-height: 420px;
    }
}

[data-testid="stHeader"] {
    display: none !important;
    height: 0 !important;
    min-height: 0 !important;
}
</style>
<script>
(function() {
  function fix() {
    var mc = document.querySelector('[data-testid="stMainBlockContainer"]');
    var hd = document.querySelector('[data-testid="stHeader"]');
    var hBlocks = document.querySelectorAll('[data-testid="stHorizontalBlock"]');
    if (mc) { mc.style.setProperty('padding-top','0','important'); mc.style.setProperty('margin-top','0','important'); }
    if (hd) { hd.style.setProperty('display','none','important'); hd.style.setProperty('height','0','important'); }
    hBlocks.forEach(function(b) {
      b.style.setProperty('margin-top','0px','important');
      b.style.setProperty('padding-top','0','important');
    });
  }
  fix();
  new MutationObserver(fix).observe(document.documentElement, {subtree:true, attributes:true, attributeFilter:['style','class']});
})();
</script>
"""


def stretch_kwargs():
    return {"width": "stretch"}


# ─── Feedback modal ─────────────────────────────────────────────────────────
FEEDBACK_MODAL_HTML = """
<div id="fb-overlay" class="fb-overlay" style="display:none;">
  <div class="fb-modal">
    <div class="fb-title" id="fb-modal-title">ช่วยให้คะแนนประสบการณ์ใช้งาน</div>
    <div class="fb-sub" id="fb-modal-sub">ความคิดเห็นของคุณช่วยให้ RecAipt ดีขึ้นได้จริงๆ ครับ</div>

    <div class="fb-star-row" id="fb-stars">
      <span class="fb-star" data-v="1">★</span>
      <span class="fb-star" data-v="2">★</span>
      <span class="fb-star" data-v="3">★</span>
      <span class="fb-star" data-v="4">★</span>
      <span class="fb-star" data-v="5">★</span>
    </div>

    <div class="fb-chip-row" id="fb-chips"></div>

    <textarea class="fb-textarea" id="fb-text" placeholder="แนะนำเพิ่มเติม หรืออธิบายปัญหาที่พบ (ไม่บังคับ)"></textarea>

    <div class="fb-btn-row">
      <button class="fb-btn" onclick="closeFeedback()">ข้ามไปก่อน</button>
      <button class="fb-btn primary" onclick="submitFeedback()">ส่งความคิดเห็น</button>
    </div>
  </div>
</div>

<script>
(function() {
  const CHIP_SETS = {
    new_receipt: [
      "ใช้ง่ายดี", "OCR ถูกต้อง", "OCR ผิดพลาด",
      "ช้าเกินไป", "ข้อมูลหายบางส่วน", "อยากได้ฟีเจอร์เพิ่ม"
    ]
  };

  let _starVal = 0;
  let _trigger = "new_receipt";
  const _selectedChips = new Set();

  window.openFeedback = function(trigger) {
    _trigger = trigger || "new_receipt";
    _starVal = 0;
    _selectedChips.clear();

    document.querySelectorAll(".fb-star").forEach(s => s.classList.remove("lit"));

    const chipRow = document.getElementById("fb-chips");
    chipRow.innerHTML = "";
    (CHIP_SETS[_trigger] || CHIP_SETS.new_receipt).forEach(label => {
      const chip = document.createElement("span");
      chip.className = "fb-chip";
      chip.textContent = label;
      chip.onclick = () => {
        if (_selectedChips.has(label)) {
          _selectedChips.delete(label);
          chip.classList.remove("selected");
        } else {
          _selectedChips.add(label);
          chip.classList.add("selected");
        }
      };
      chipRow.appendChild(chip);
    });

    document.getElementById("fb-text").value = "";

    document.getElementById("fb-modal-title").textContent = "เริ่มใบใหม่แล้ว — ช่วยให้คะแนนด้วยนะ";
    document.getElementById("fb-modal-sub").textContent = "ความคิดเห็นของคุณช่วยให้ RecAipt ดีขึ้นได้จริงๆ ครับ";

    document.getElementById("fb-overlay").style.display = "flex";
  };

  window.closeFeedback = function() {
    document.getElementById("fb-overlay").style.display = "none";
  };

  window.submitFeedback = function() {
    const payload = {
      trigger: _trigger,
      rating: _starVal,
      chips: Array.from(_selectedChips),
      comment: document.getElementById("fb-text").value.trim(),
      ts: new Date().toISOString()
    };
    try {
      const prev = JSON.parse(localStorage.getItem("recaipt_feedback") || "[]");
      prev.push(payload);
      localStorage.setItem("recaipt_feedback", JSON.stringify(prev));
    } catch(e) {}
    closeFeedback();
    showToast("ขอบคุณสำหรับความคิดเห็น! 🙏");
  };

  document.addEventListener("DOMContentLoaded", function() {
    hookStars();
  });
  hookStars();

  function hookStars() {
    document.querySelectorAll(".fb-star").forEach(star => {
      star.addEventListener("click", function() {
        _starVal = parseInt(this.dataset.v);
        document.querySelectorAll(".fb-star").forEach(s => {
          s.classList.toggle("lit", parseInt(s.dataset.v) <= _starVal);
        });
      });
    });
  }

  window.showToast = function(msg) {
    const t = document.createElement("div");
    t.textContent = msg;
    t.style.cssText = `
      position:fixed;bottom:28px;left:50%;transform:translateX(-50%);
      background:#2F6F73;color:#fff;padding:10px 22px;border-radius:999px;
      font-size:0.98rem;font-weight:650;z-index:99999;
      box-shadow:0 8px 24px rgba(15,40,48,0.22);
      animation:fadeUp 0.25s ease;
    `;
    document.body.appendChild(t);
    setTimeout(() => t.remove(), 2800);
  };
})();
</script>
<style>
@keyframes fadeUp {
  from { opacity:0; transform:translateX(-50%) translateY(12px); }
  to   { opacity:1; transform:translateX(-50%) translateY(0); }
}
</style>
"""


def safe_float(value, default=0.0):
    try:
        if value in (None, ""):
            return default
        parsed = float(str(value).replace(",", "").strip())
        return parsed if math.isfinite(parsed) else default
    except (TypeError, ValueError):
        return default


def normalize_date(value):
    if not value:
        return ""
    text = str(value).strip()
    numeric = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", text)
    iso = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", text)
    if iso:
        year, month, day = map(int, iso.groups())
        if year > 2500:
            year -= 543
        return f"{year:04d}-{month:02d}-{day:02d}"
    if numeric:
        day, month, year = map(int, numeric.groups())
        if year > 2500:
            year -= 543
        return f"{year:04d}-{month:02d}-{day:02d}"
    return text


def get_nested(data, path, default=""):
    current = data or {}
    for key in path:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
    return default if current in (None, "") else current


def derive_vat_values(grand_total, vat_rate=7.0, tax_included=True):
    _total = safe_float(grand_total)
    _rate = safe_float(vat_rate, 7.0)
    if _total <= 0 or _rate <= 0:
        return 0.0, 0.0
    if tax_included:
        before_tax = _total / (1 + (_rate / 100))
        vat_val = _total - before_tax
        return round(before_tax, 2), round(vat_val, 2)
    vat_val = _total * (_rate / 100)
    return round(_total, 2), round(vat_val, 2)


def get_summary(data):
    summary = data.get("summary") if isinstance(data, dict) else {}
    return summary if isinstance(summary, dict) else {}


def get_document_header(data):
    header = data.get("document_header") if isinstance(data, dict) else {}
    return header if isinstance(header, dict) else {}


def get_merchant_and_buyer(data):
    mb = data.get("merchant_and_buyer") if isinstance(data, dict) else {}
    return mb if isinstance(mb, dict) else {}


def normalize_result(data):
    data = data or {}
    header = get_document_header(data)
    mb = get_merchant_and_buyer(data)
    summary = get_summary(data)

    items = []
    for item in data.get("line_items", data.get("items", [])) or []:
        if not isinstance(item, dict):
            continue
        qty = safe_float(item.get("quantity", item.get("qty", 1)), 1.0)
        unit_price = safe_float(item.get("unit_price", 0.0))
        total_price = safe_float(
            item.get("total_price", item.get("subtotal", item.get("amount", qty * unit_price)))
        )
        items.append(
            {
                "item_name": item.get("item_name", item.get("item_description", item.get("name"))) or None,
                "quantity": qty if qty != 0 else None,
                "unit_price": unit_price if unit_price != 0 else None,
                "total_price": total_price if total_price != 0 else None,
            }
        )

    if not items:
        items = [{"item_name": None, "quantity": None, "unit_price": None, "total_price": None}]

    # GOLDEN_SCHEMA ใหม่ไม่มี total_before_discount แล้ว แต่มี vat_amount แทน
    # total_before_discount (ยอดรวมสินค้าก่อนหักส่วนลด) ยังใช้ภายในสำหรับคำนวณ
    # แต่คำนวณจาก subtotal + vat_amount แทนที่จะรับมาตรงๆ จาก LLM
    subtotal = safe_float(summary.get("subtotal", data.get("amount_before_tax", 0.0)))
    vat_amount = safe_float(summary.get("vat_amount", data.get("vat_amount", 0.0)))
    discount_total = safe_float(summary.get("total_discount", data.get("discount_total", 0.0)))
    net_amount = safe_float(summary.get("net_amount", data.get("net_payable", 0.0)))

    total_before_discount = round(subtotal + vat_amount, 2)

    # ป้องกันกรณี LLM ดึงยอดมาเป็น "ยอดสุทธิ" ที่หักส่วนลด/แต้ม/คูปองออกแล้ว
    # ซึ่งทำให้ฐาน VAT ต่ำกว่ามูลค่าสินค้าจริง: ถ้าผลรวมรายการสินค้า (line_items) มากกว่า
    # total_before_discount ที่คำนวณได้ ให้ใช้ผลรวมรายการสินค้าแทน แล้วคำนวณ subtotal/vat_amount ใหม่
    items_sum = round(sum(safe_float(it.get("total_price")) for it in items), 2)
    if items_sum > 0 and items_sum > total_before_discount + 0.01:
        total_before_discount = items_sum
        subtotal, vat_amount = derive_vat_values(total_before_discount, 7.0, tax_included=True)

    if total_before_discount > 0 and subtotal <= 0:
        subtotal, vat_amount = derive_vat_values(total_before_discount, 7.0, tax_included=True)

    # derive net_amount ถ้า LLM ไม่ส่งมา
    if net_amount <= 0 and total_before_discount > 0:
        net_amount = max(total_before_discount - discount_total, 0.0)
        net_amount = round(net_amount, 2)

    return {
        "document_header": {
            "document_type": normalize_doc_type(header.get("document_type", data.get("document_type"))),
            "document_date": normalize_date(header.get("document_date", data.get("document_date"))) or None,
            "document_number": header.get("document_number", data.get("document_number")) or None,
        },
        "merchant_and_buyer": {
            "merchant_name": mb.get("merchant_name", data.get("store_name")) or None,
            "buyer_name": mb.get("buyer_name") or None,
            "merchant_tax_id": mb.get("merchant_tax_id", data.get("tax_id")) or None,
            "buyer_tax_id": mb.get("buyer_tax_id") or None,
        },
        "line_items": items,
        "summary": {
            "subtotal": subtotal if subtotal != 0 else None,
            "total_discount": round(discount_total, 2) if discount_total != 0 else None,
            "vat_amount": round(vat_amount, 2) if vat_amount != 0 else None,
            "net_amount": round(net_amount, 2)
            if net_amount != total_before_discount and net_amount != 0
            else None,
        },
    }


def build_export_payload(values, edited_items):
    if hasattr(edited_items, "to_dict"):
        edited_items = edited_items.to_dict("records")

    items = []
    for item in edited_items:
        name = str(item.get("item_name") or "").strip()
        qty = safe_float(item.get("quantity"), 0.0)
        unit_price = safe_float(item.get("unit_price"), 0.0)
        total_price = safe_float(item.get("total_price"), qty * unit_price)
        if not name and total_price == 0:
            continue
        items.append(
            {
                "item_name": name or None,
                "quantity": qty if qty != 0 else None,
                "unit_price": unit_price if unit_price != 0 else None,
                "total_price": total_price if total_price != 0 else None,
            }
        )

    if not items:
        items = [{"item_name": None, "quantity": None, "unit_price": None, "total_price": None}]

    return {
        "document_header": {
            "document_type": normalize_doc_type(values["document_type"]) or DOC_TYPE_DEFAULT,
            "document_date": normalize_date(values["document_date"]) or None,
            "document_number": values["document_number"] or None,
        },
        "merchant_and_buyer": {
            "merchant_name": values["merchant_name"] or None,
            "buyer_name": values["buyer_name"] or None,
            "merchant_tax_id": values["merchant_tax_id"] or None,
            "buyer_tax_id": values["buyer_tax_id"] or None,
        },
        "line_items": items,
        "summary": {
            "subtotal": values["subtotal"] if values["subtotal"] != 0 else None,
            "total_discount": values.get("total_discount") if values.get("total_discount", 0) != 0 else None,
            "vat_amount": values.get("vat_amount") if values.get("vat_amount", 0) != 0 else None,
            "net_amount": values.get("net_amount")
            if values.get("net_amount", 0) not in (0, values.get("total_before_discount"))
            else None,
        },
    }


def build_receipt_csv(payload):
    output = StringIO()
    fieldnames = [
        "document_type", "document_number", "document_date",
        "merchant_name", "merchant_tax_id", "buyer_name", "buyer_tax_id",
        "subtotal", "vat_amount",
        "total_discount", "net_amount",
        "item_no", "item_name", "quantity", "unit_price", "item_total_price",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    header = get_document_header(payload)
    mb = get_merchant_and_buyer(payload)
    totals = get_summary(payload)
    items = payload.get("line_items", []) or [{}]
    for index, item in enumerate(items, start=1):
        writer.writerow(
            {
                "document_type": header.get("document_type", ""),
                "document_number": header.get("document_number", ""),
                "document_date": header.get("document_date", ""),
                "merchant_name": mb.get("merchant_name", ""),
                "merchant_tax_id": mb.get("merchant_tax_id", ""),
                "buyer_name": mb.get("buyer_name", ""),
                "buyer_tax_id": mb.get("buyer_tax_id", ""),
                "subtotal": f"{safe_float(totals.get('subtotal')):.2f}",
                "vat_amount": f"{safe_float(totals.get('vat_amount')):.2f}",
                "total_discount": f"{safe_float(totals.get('total_discount')):.2f}",
                "net_amount": f"{safe_float(totals.get('net_amount')) or (safe_float(totals.get('subtotal')) + safe_float(totals.get('vat_amount'))):.2f}",
                "item_no": index,
                "item_name": item.get("item_name", ""),
                "quantity": f"{safe_float(item.get('quantity'), 1.0):.2f}",
                "unit_price": f"{safe_float(item.get('unit_price')):.2f}",
                "item_total_price": f"{safe_float(item.get('total_price')):.2f}",
            }
        )
    return output.getvalue()


def build_receipt_excel(payload):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("กรุณาติดตั้ง openpyxl ก่อน: pip install openpyxl")
    from datetime import datetime
    from io import BytesIO

    ACCENT = "00A8CC"
    ACCENT_LIGHT = "EAF6FB"
    WARN = "FFF3D8"
    OK = "E8F2E7"
    WHITE = "FFFFFF"
    INK = "1F2937"
    MUTED = "5B6B73"

    def header_style(cell, bg=ACCENT, fg="FFFFFF", bold=True):
        cell.font = Font(bold=bold, color=fg, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def data_style(cell, bg=WHITE, align="left", bold=False):
        cell.font = Font(color=INK, size=10, bold=bold)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        thin = Side(style="thin", color="DDDDDD")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "รายการสินค้า"
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 32

    header = get_document_header(payload)
    mb = get_merchant_and_buyer(payload)
    totals = get_summary(payload)

    ws1.merge_cells("A1:J1")
    title_cell = ws1["A1"]
    title_cell.value = f"RecAipt — {mb.get('merchant_name', '-')}  |  เลขที่: {header.get('document_number', '-')}  |  วันที่: {header.get('document_date', '-')}"
    title_cell.font = Font(bold=True, color="FFFFFF", size=12)
    title_cell.fill = PatternFill("solid", fgColor=ACCENT)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 36

    info_rows = [
        ("ผู้ขาย", mb.get("merchant_name", "-"), "เลขผู้เสียภาษีผู้ขาย", mb.get("merchant_tax_id", "-")),
        ("ผู้ซื้อ", mb.get("buyer_name", "-") or "-", "เลขผู้เสียภาษีผู้ซื้อ", mb.get("buyer_tax_id", "-") or "-"),
        ("ประเภทเอกสาร", header.get("document_type", "-"), "วันที่เอกสาร", header.get("document_date", "-") or "-"),
    ]
    for r_idx, (l1, v1, l2, v2) in enumerate(info_rows, start=2):
        ws1.merge_cells(f"A{r_idx}:B{r_idx}")
        ws1.merge_cells(f"C{r_idx}:E{r_idx}")
        ws1.merge_cells(f"F{r_idx}:G{r_idx}")
        ws1.merge_cells(f"H{r_idx}:J{r_idx}")
        for col, val, bg, bold in [(1, l1, ACCENT_LIGHT, True), (3, v1, WHITE, False),
                                   (6, l2, ACCENT_LIGHT, True), (8, v2, WHITE, False)]:
            c = ws1.cell(row=r_idx, column=col, value=val)
            data_style(c, bg=bg, bold=bold)
        ws1.row_dimensions[r_idx].height = 20

    ws1.row_dimensions[6].height = 8

    item_headers = ["#", "รายการสินค้า / บริการ", "จำนวน", "หน่วย", "ราคา/หน่วย",
                    "ยอดรวมรายการ", "สถานะ OCR", "", "", ""]
    col_widths = [5, 38, 9, 8, 14,
                  16, 14, 0, 0, 0]
    ITEM_HEADER_ROW = 7
    for col_i, (hdr, w) in enumerate(zip(item_headers, col_widths), start=1):
        c = ws1.cell(row=ITEM_HEADER_ROW, column=col_i, value=hdr)
        header_style(c)
        if w:
            ws1.column_dimensions[get_column_letter(col_i)].width = w
    ws1.row_dimensions[ITEM_HEADER_ROW].height = 28

    items = payload.get("line_items", []) or []
    for i, item in enumerate(items, start=1):
        row = ITEM_HEADER_ROW + i
        bg = WHITE if i % 2 == 1 else "EAF6FB"
        values = [
            i, item.get("item_name", "-"), safe_float(item.get("quantity"), 1.0), "หน่วย",
            safe_float(item.get("unit_price")), safe_float(item.get("total_price")),
            item.get("สถานะ", "-") if "สถานะ" in item else "-",
        ]
        aligns = ["center", "left", "center", "center", "right", "right", "center"]
        for col_i, (val, aln) in enumerate(zip(values, aligns), start=1):
            c = ws1.cell(row=row, column=col_i, value=val)
            data_style(c, bg=bg, align=aln)
            if col_i in (5, 6):
                c.number_format = '#,##0.00'
            if col_i == 3:
                c.number_format = '#,##0.00'
        ws1.row_dimensions[row].height = 22

    sum_row = ITEM_HEADER_ROW + len(items) + 2
    _subtotal_xl = safe_float(totals.get("subtotal"))
    _vat_xl = safe_float(totals.get("vat_amount"))
    _total_before_discount_xl = round(_subtotal_xl + _vat_xl, 2)
    _discount_val_xl = safe_float(totals.get("total_discount"))
    _net_amount_xl = safe_float(totals.get("net_amount")) or _total_before_discount_xl
    summary = [
        ("ยอดก่อนภาษี (Subtotal)", _subtotal_xl, ACCENT_LIGHT),
        ("ภาษีมูลค่าเพิ่ม (VAT)", _vat_xl, ACCENT_LIGHT),
        ("ยอดรวมสินค้า (Total Before Discount)", _total_before_discount_xl, OK),
    ]
    if _discount_val_xl > 0:
        summary.append(("ส่วนลดรวม (Discount)", -_discount_val_xl, "FFF3D8"))
        summary.append(("ยอดที่จ่ายจริง (Net Amount)", _net_amount_xl, "D6EFD8"))
    for offset, (label, value, bg) in enumerate(summary):
        r = sum_row + offset
        ws1.merge_cells(f"A{r}:E{r}")
        lc = ws1.cell(row=r, column=1, value=label)
        data_style(lc, bg=bg, align="right", bold=True)
        vc = ws1.cell(row=r, column=6, value=value)
        data_style(vc, bg=bg, align="right", bold=True)
        vc.number_format = '#,##0.00'
        ws1.row_dimensions[r].height = 24

    note_row = sum_row + len(summary) + 2
    ws1.merge_cells(f"A{note_row}:E{note_row}")
    note = ws1[f"A{note_row}"]
    note.value = f"ไฟล์ต้นฉบับ: {payload.get('_filename', st.session_state.get('file_name', '-'))}  |  ส่งออกเมื่อ: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    note.font = Font(color=MUTED, italic=True, size=9)
    note.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[note_row].height = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def review_flags(data):
    flags = []
    if not get_nested(data, ["merchant_and_buyer", "merchant_name"]):
        flags.append("ยังไม่พบชื่อร้านค้า ควรตรวจจากหัวใบเสร็จ")
    if not get_nested(data, ["document_header", "document_date"]):
        flags.append("ยังไม่พบวันที่เอกสาร")
    if not get_nested(data, ["document_header", "document_number"]):
        flags.append("ยังไม่พบเลขที่เอกสาร")
    _flag_summary = get_summary(data)
    _flag_total = safe_float(_flag_summary.get("net_amount")) or (
        safe_float(_flag_summary.get("subtotal")) + safe_float(_flag_summary.get("vat_amount"))
    )
    if _flag_total <= 0:
        flags.append("ยอดรวมสุทธิยังเป็น 0 หรือไม่ชัดเจน")
    if not data.get("line_items"):
        flags.append("ยังไม่พบรายการสินค้า")
    return flags


def image_to_data_uri(image_np, rotation=0):
    image = image_np
    if rotation == 90:
        image = cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)
    elif rotation == 180:
        image = cv2.rotate(image, cv2.ROTATE_180)
    elif rotation == 270:
        image = cv2.rotate(image, cv2.ROTATE_90_COUNTERCLOCKWISE)
    if len(image.shape) == 2:
        ok, encoded = cv2.imencode(".png", image)
    else:
        rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        ok, encoded = cv2.imencode(".png", rgb)
    if not ok:
        return ""
    return "data:image/png;base64," + base64.b64encode(encoded.tobytes()).decode("ascii")


def copy_button(label, text, key):
    payload = json.dumps(str(text), ensure_ascii=False)
    components.html(
        f"""
        <style>
        body {{
            margin: 0;
            padding: 0;
            overflow: hidden;
            background: transparent;
        }}
        #copy-{key} {{
            width: 100%;
            height: 44px;
            border-radius: 8px;
            border: 1px solid #D8D1C3;
            background: #FFFDF8;
            color: #1F2A2A;
            font-size: 16px;
            font-weight: 720;
            cursor: pointer;
            box-sizing: border-box;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: background 0.15s, border-color 0.15s, color 0.15s;
        }}
        #copy-{key}:hover {{
            border-color: #2F6F73;
            color: #285E62;
            background: #EEF5F3;
        }}
        </style>
        <button id="copy-{key}">{label}</button>
        <script>
        const btn = document.getElementById("copy-{key}");
        btn.addEventListener("click", async () => {{
            await navigator.clipboard.writeText({payload});
            btn.innerText = "คัดลอกแล้ว";
            setTimeout(() => btn.innerText = {json.dumps(label, ensure_ascii=False)}, 1200);
        }});
        </script>
        """,
        height=45,
    )


def reset_app():
    for key in list(st.session_state.keys()):
        del st.session_state[key]


def run_pipeline(uploaded_file):
    file_bytes = uploaded_file.read()
    file_name = uploaded_file.name

    progress = st.progress(0, text="กำลังเตรียมภาพเอกสาร")
    img = load_image_or_pdf(file_bytes, file_name)
    if img is None:
        st.error("ไฟล์เอกสารเสียหาย หรือระบบไม่รองรับรูปแบบภาพนี้")
        st.stop()

    deskewed_img = deskew_image(img)
    processed_img = process_clahe(deskewed_img)
    progress.progress(35, text="กำลังส่งภาพเข้า Typhoon OCR")

    raw_text = run_typhoon_ocr(processed_img)
    if "[ERROR]" in raw_text or not raw_text.strip():
        progress.empty()
        st.error(raw_text if raw_text.strip() else "ระบบไม่สามารถถอดข้อความจากใบเสร็จนี้ได้")
        st.stop()

    formatted_text = format_ocr_text(clean_ocr_text(raw_text))

    progress.progress(70, text="กำลังสกัดข้อมูลด้วย Typhoon LLM")
    extracted = call_typhoon_llm(formatted_text)
    progress.progress(100, text="เสร็จสิ้น")
    progress.empty()

    if isinstance(extracted, dict) and "error" in extracted:
        st.error(extracted["error"])
        st.stop()

    st.session_state["processed_img"] = processed_img
    st.session_state["raw_text"] = raw_text
    st.session_state["formatted_ocr_text"] = formatted_text
    st.session_state["result_json"] = normalize_result(extracted)
    st.session_state["file_name"] = file_name


def render_topbar(result=None):
    import html as _html

    if result is not None:
        flags = review_flags(result)
        review_badge = '⚠ รอตรวจสอบบางช่อง' if flags else '✓ ข้อมูลครบถ้วน'
        raw_filename = st.session_state.get("file_name", "")
        filename = _html.escape(str(raw_filename))
        file_badge = f'📄 {filename}' if filename else ""

        status_html = f'<span class="badge badge-ok">OCR + LLM</span>'
        status_html += f'<span class="badge badge-warn">{review_badge}</span>'
        if file_badge:
            status_html += f'<span class="badge badge-bad">{file_badge}</span>'
    else:
        status_html = '<span class="badge badge-ok">OCR + LLM Pipeline พร้อมใช้งาน</span>'

    st.markdown(CSS_STYLES, unsafe_allow_html=True)
    st.markdown(
        f'<div class="app-topbar">'
        f'<div class="topbar-grid">'
        f'<div class="topbar-left">'
        f'<span class="brand">RecAipt</span>'
        f'<div class="topbar-divider"></div>'
        f'<div class="status-row">{status_html}</div>'
        f'</div>'
        f'</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


if "result_json" not in st.session_state:
    render_topbar()
    st.markdown(
        """
        <div class="upload-page-shell">
          <div class="upload-wrap">
            <div class="upload-title">อัปโหลดใบเสร็จเพื่อเริ่มตรวจสอบ</div>
            <div class="upload-subtitle">
              รองรับไฟล์ JPG, PNG และ PDF หน้าแรก ระบบจะปรับภาพด้วย Method 4 แล้วส่งเข้า Typhoon OCR และ Typhoon LLM
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    uploaded_file = st.file_uploader(
        "เลือกไฟล์ใบเสร็จ",
        type=["jpg", "jpeg", "png", "pdf"],
        label_visibility="collapsed",
    )

    if uploaded_file is not None:
        run_pipeline(uploaded_file)
        st.rerun()

    st.markdown(
        """
        <div class="upload-tip">
          <div class="soft-callout">
            คำแนะนำ: วางไฟล์ในพื้นที่อัปโหลดหรือลากมาวางได้โดยตรง หลังประมวลผลแล้วให้ตรวจช่องสีเหลืองก่อน
            จากนั้นใช้ Tab เพื่อเลื่อนไปยังช่องถัดไป และส่งออกข้อมูลเมื่อยืนยันครบถ้วน
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()

# ─── บล็อกกรณีประมวลผลหน้าต่างตรวจสอบข้อมูล ───
result_json = st.session_state["result_json"]
render_topbar(result_json)

left, right = st.columns([0.95, 1.05], gap="medium")

with left:
    st.markdown(
        """
        <div class="main-pane-start">
          <div class="pane">
            <div class="pane-head">
              <div class="pane-title">ภาพใบเสร็จหลังปรับคมชัด</div>
              <div class="pane-note">ใช้ซูมและหมุนภาพเพื่อเทียบตัวเลขกับข้อมูลด้านขวา ลดการสลับหน้าจอไปมา</div>
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="zoom-control-row">', unsafe_allow_html=True)
    zoom_col, rotate_col = st.columns([1, 1])

    with zoom_col:
        zoom = st.slider("", min_value=80, max_value=190, value=110, step=10)
    with rotate_col:
        rotation = st.selectbox("", [0, 90, 180, 270], format_func=lambda x: f"{x} องศา")

    st.markdown("</div>", unsafe_allow_html=True)

    data_uri = image_to_data_uri(st.session_state["processed_img"], rotation)
    st.markdown(
        f"""
        <div class="image-scroll">
          <img class="receipt-image" src="{data_uri}" style="width:{zoom}%;" alt="processed receipt">
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="ocr-expander-spacer"></div>', unsafe_allow_html=True)

    with st.expander("ข้อความ OCR ดิบ"):
        st.caption("แสดงแบบจัดบรรทัดเพื่ออ่านง่ายขึ้น โดยยังคงข้อความ OCR ต้นทางไว้สำหรับตรวจสอบ")
        st.code(st.session_state.get("formatted_ocr_text", st.session_state.get("raw_text", "")), language="text")
        copy_button(
            "คัดลอกข้อความ OCR",
            st.session_state.get("formatted_ocr_text", st.session_state.get("raw_text", "")),
            "raw-ocr",
        )

with right:
    st.markdown(
        """
        <div class="main-pane-start">
          <div class="pane">
            <div class="pane-head">
              <div class="pane-title">ข้อมูลดิจิทัลสำหรับตรวจสอบ</div>
              <div class="pane-note">แก้ไขได้ทุกช่อง ตารางรายการรองรับการใช้คีย์บอร์ดและการเพิ่มแถว</div>
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    flags = review_flags(result_json)
    st.markdown('<div class="section-band">จุดที่ควรตรวจสอบก่อน</div>', unsafe_allow_html=True)

    if flags:
        st.markdown(
            "<ul class='review-list'>" +
            "".join([f"<li class='needs-check'>{flag}</li>" for flag in flags]) +
            "</ul>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            "<ul class='review-list'><li class='looks-ok'>ข้อมูลครบถ้วนแล้ว ตรวจทานรายการสินค้าและยอดรวมอีกครั้งก่อนส่งออก</li></ul>",
            unsafe_allow_html=True,
        )

    _ft = get_summary(result_json)
    _total_before_discount_val = safe_float(_ft.get("subtotal")) + safe_float(_ft.get("vat_amount"))
    _discount_val = safe_float(_ft.get("total_discount"))
    _net_amount_val = safe_float(_ft.get("net_amount")) or _total_before_discount_val
    _has_discount = _discount_val > 0
    st.markdown(
        f"""
        <div class="metric-grid">
          <div class="metric">
            <div class="metric-label">ชื่อร้านค้า</div>
            <div class="metric-value">{get_nested(result_json, ["merchant_and_buyer", "merchant_name"], "-") or "-"}</div>
          </div>
          <div class="metric">
            <div class="metric-label">วันที่เอกสาร</div>
            <div class="metric-value">{get_nested(result_json, ["document_header", "document_date"], "-") or "-"}</div>
          </div>
          <div class="metric">
            <div class="metric-label">{"ยอดจ่ายจริง (หลังหักส่วนลด)" if _has_discount else "ยอดรวมสินค้า"}</div>
            <div class="metric-value" style="{"color:#c0392b" if _has_discount else ""}">{_net_amount_val:,.2f}</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.form("receipt_verification_form", clear_on_submit=False):
        st.markdown('<div class="section-band">หัวเอกสาร</div>', unsafe_allow_html=True)
        doc_col1, doc_col2 = st.columns([1, 1])
        with doc_col1:
            _current_doc_type = normalize_doc_type(get_nested(result_json, ["document_header", "document_type"]))
            document_type = st.selectbox(
                "ประเภทเอกสาร",
                options=DOC_TYPES,
                index=DOC_TYPES.index(_current_doc_type),
            )
            document_number = st.text_input(
                "เลขที่เอกสาร", value=get_nested(result_json, ["document_header", "document_number"], "")
            )
        with doc_col2:
            document_date = st.text_input(
                "วันที่เอกสาร (YYYY-MM-DD)", value=get_nested(result_json, ["document_header", "document_date"], "")
            )

        st.markdown('<div class="section-band">ผู้ขายและผู้ซื้อ</div>', unsafe_allow_html=True)
        seller_col, buyer_col = st.columns([1, 1])
        with seller_col:
            merchant_name = st.text_input(
                "ชื่อร้านค้า / ผู้ขาย", value=get_nested(result_json, ["merchant_and_buyer", "merchant_name"])
            )
            merchant_tax_id = st.text_input(
                "เลขผู้เสียภาษีผู้ขาย", value=get_nested(result_json, ["merchant_and_buyer", "merchant_tax_id"])
            )
        with buyer_col:
            buyer_name = st.text_input(
                "ชื่อผู้ซื้อ", value=get_nested(result_json, ["merchant_and_buyer", "buyer_name"])
            )
            buyer_tax_id = st.text_input(
                "เลขผู้เสียภาษีผู้ซื้อ", value=get_nested(result_json, ["merchant_and_buyer", "buyer_tax_id"])
            )

        st.markdown('<div class="section-band">รายการสินค้าและบริการ</div>', unsafe_allow_html=True)
        edited_items = st.data_editor(
            result_json.get("line_items", []),
            num_rows="dynamic",
            **stretch_kwargs(),
            hide_index=True,
            column_config={
                "item_name": st.column_config.TextColumn("รายการ", width="large"),
                "quantity": st.column_config.NumberColumn("จำนวน", min_value=0.0, step=1.0, format="%.2f"),
                "unit_price": st.column_config.NumberColumn("ราคาต่อหน่วย", min_value=0.0, step=0.25, format="%.2f"),
                "total_price": st.column_config.NumberColumn("รวม", min_value=0.0, step=0.25, format="%.2f"),
            },
            key="items_editor",
        )

        st.markdown('<div class="section-band">สรุปยอด</div>', unsafe_allow_html=True)
        summary_totals = get_summary(result_json)
        tax_mode = st.selectbox(
            "รูปแบบ VAT",
            ["VAT รวมในราคาสินค้า", "VAT แยกบวกเพิ่ม"],
            index=0 if st.session_state.get("tax_included", True) else 1,
            help="ใบเสร็จร้านค้าปลีก เช่น 7-Eleven มักเป็น VAT รวมในราคาสินค้า ไม่ควรบวก VAT ซ้ำอีกครั้ง",
        )
        tax_included = tax_mode == "VAT รวมในราคาสินค้า"
        st.session_state["tax_included"] = tax_included
        _existing_total_before_discount = safe_float(summary_totals.get("subtotal")) + safe_float(
            summary_totals.get("vat_amount")
        )

        total_col1, total_col2 = st.columns([1, 1])
        with total_col1:
            suggested_subtotal, suggested_vat = derive_vat_values(
                _existing_total_before_discount,
                7.0,
                tax_included=tax_included,
            )
            subtotal = st.number_input(
                "ยอดก่อนภาษี",
                value=safe_float(summary_totals.get("subtotal")) or suggested_subtotal,
                min_value=0.0,
                step=0.25,
                format="%.2f",
            )
        with total_col2:
            vat_amount = st.number_input(
                "ภาษีมูลค่าเพิ่ม (VAT)",
                value=safe_float(summary_totals.get("vat_amount")) or suggested_vat,
                min_value=0.0,
                step=0.25,
                format="%.2f",
            )

        total_before_discount = round(subtotal + vat_amount, 2)

        # Row ส่วนลด + ยอดจ่ายจริง
        discount_col1, discount_col2 = st.columns([1, 1])
        with discount_col1:
            total_discount = st.number_input(
                "ส่วนลดรวม (M-Stamp / คูปอง / แต้ม)",
                value=safe_float(summary_totals.get("total_discount")),
                min_value=0.0,
                step=0.25,
                format="%.2f",
            )
        with discount_col2:
            _derived_net = round(max(total_before_discount - total_discount, 0.0), 2)
            net_amount = st.number_input(
                "ยอดที่จ่ายจริง (หลังหักส่วนลด)",
                value=safe_float(summary_totals.get("net_amount")) or _derived_net or total_before_discount,
                min_value=0.0,
                step=0.25,
                format="%.2f",
            )

        st.markdown(
            """
            <div class="action-pad">
              <div class="tiny-note">
                ใช้ <span class="kbd">Tab</span> เพื่อย้ายช่องกรอกข้อมูล และตรวจช่องที่ยังเป็นสีเตือนก่อนบันทึก
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        save_local = st.form_submit_button("ยืนยันข้อมูลบนหน้านี้", type="primary", **stretch_kwargs())

current_values = {
    "document_type": document_type,
    "document_number": document_number,
    "document_date": document_date,
    "merchant_name": merchant_name,
    "merchant_tax_id": merchant_tax_id,
    "buyer_name": buyer_name,
    "buyer_tax_id": buyer_tax_id,
    "subtotal": subtotal,
    "total_before_discount": total_before_discount,
    "vat_amount": vat_amount,
    "total_discount": total_discount,
    "net_amount": net_amount,
}
export_payload = build_export_payload(current_values, edited_items)

if save_local:
    st.session_state["result_json"] = normalize_result(export_payload)
    st.success("บันทึกค่าที่ตรวจสอบไว้บนหน้านี้แล้ว")
    st.rerun()

# ─── ชุดปุ่มดำเนินการและปุ่มดาวน์โหลด ───
with right:
    st.markdown('<div class="section-band">ส่งออกข้อมูล</div>', unsafe_allow_html=True)
    output_name = (get_nested(export_payload, ["document_header", "document_number"]) or "receipt").replace("/", "-")

    dl_col1, dl_col2, dl_col3 = st.columns([1, 1, 1])
    with dl_col1:
        st.download_button(
            "⬇ CSV",
            data=build_receipt_csv(export_payload).encode("utf-8-sig"),
            file_name=f"{output_name}.csv",
            mime="text/csv",
            **stretch_kwargs(),
        )
    with dl_col2:
        st.download_button(
            "⬇ JSON",
            data=json.dumps(export_payload, ensure_ascii=False, indent=2).encode("utf-8"),
            file_name=f"{output_name}.json",
            mime="application/json",
            **stretch_kwargs(),
        )
    with dl_col3:
        st.download_button(
            "⬇ Excel",
            data=build_receipt_excel(export_payload),
            file_name=f"{output_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            **stretch_kwargs(),
        )

    cp_col1, cp_col2 = st.columns([1, 1])
    with cp_col1:
        copy_button("คัดลอก JSON", json.dumps(export_payload, ensure_ascii=False, indent=2), "json")
    with cp_col2:
        _copy_ft = get_summary(export_payload)
        _copy_net = safe_float(_copy_ft.get("net_amount")) or (
            safe_float(_copy_ft.get("subtotal")) + safe_float(_copy_ft.get("vat_amount"))
        )
        copy_button(
            "คัดลอกยอดที่จ่ายจริง",
            f"{_copy_net:,.2f}",
            "net-payable",
        )

    st.markdown('<div class="section-band">ดำเนินการต่อ</div>', unsafe_allow_html=True)
    if st.button("🔄 สแกนใบเสร็จใหม่", key="cloud_action_new_scan", **stretch_kwargs()):
        reset_app()
        st.rerun()

# โหลดโมดอลไว้ล่างสุด
components.html(FEEDBACK_MODAL_HTML, height=0)
